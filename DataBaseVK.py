import vk_api
from tqdm import tqdm
from openpyxl import load_workbook
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
from pyfiglet import Figlet

f=Figlet(font='slant')
print(u"\u001b[35m"+f.renderText('D a t a b a s e VK'))

tok=input('Введите свой токен ВК: ')
put=input('Укажите путь к файлу .xlsx (имя листа data!): ')
vib=int(input('Укажите режим работы(0-через друга, 1-диапазон айди): '))

if vib==1:
    left=int(input('Укажите левую границу айди: '))
    rite=int(input('Укажите правую границу айди: '))
    a=range(left, rite)
else:
    polz = input('Укажите айди друга для сбора данных: ')
print(' ')


session=vk_api.VkApi(token=tok)
vk=session.get_api()

if vib==0:
    user_id=polz
    a = session.method('friends.get', {'user_id': user_id})['items']

fn= put
wb=load_workbook(fn)
ws=wb['data']
noom=1
for i in tqdm(a):
    noom+=1
    profiles = vk.users.get(user_id=i, fields="about, military, sex, bdate, career, city, connections, contacts, counters, country, education, followers_count, friend_status, home_town, is_friend, is_no_index")
    try:
        ws['A1']='ВК айди'
        ws['A'+str(noom)]=profiles[0]['id']
    except:
        gg=1
    try:
        ws['C1']='Имя'
        ws['C' + str(noom)] = profiles[0]['first_name']
    except:
        gg=1
    try:
        ws['B1'] = 'Фамилия'
        ws['B' + str(noom)] = profiles[0]['last_name']
    except:
        gg=1
    try:
        ws['D1'] = 'Пол'
        ws['D' + str(noom)] = profiles[0]['sex']
    except:
        gg=1
    try:
        ws['E1'] = 'Дата рождения'
        ws['E' + str(noom)] = profiles[0]['bdate']
    except:
        gg=1
    try:
        ws['F1'] = 'Бан'
        ws['F' + str(noom)] = profiles[0]['deactivated']
    except:
        gg=1
    try:
        ws['G1'] = 'Скрытый пользователь'
        ws['G' + str(noom)] = profiles[0]['is_closed']
    except:
        gg=1
    try:
        ws['H1'] = 'О себе'
        ws['H' + str(noom)] = profiles[0]['about']
    except:
        gg=1
    try:
        ws['I1'] = 'Название компании, должность'
        ws['I' + str(noom)] = profiles[0]['career'][0]['company']
    except:
        gg=1
    try:
        ws['J1'] = 'Место учебы'
        ws['J' + str(noom)] = profiles[0]['university_name']
    except:
        gg=1
    try:
        ws['K1'] = 'Если служил, номер части'
        ws['K' + str(noom)] = profiles[0]['military'][0]['unit']
    except:
        gg=1
    try:
        ws['L1'] = 'skype '
        ws['L' + str(noom)] = profiles[0]['skype']
    except:
        gg=1
    try:
        ws['M1'] = 'Страна'
        ws['M' + str(noom)] = profiles[0]['country']['title']
    except:
        gg=1
    try:
        ws['N1'] = 'Город'
        ws['N' + str(noom)] = profiles[0]['city']['title']
    except:
        gg=1
    try:
        ws['O1'] = 'Телефон'
        ws['O' + str(noom)] = profiles[0]['mobile_phone']
    except:
        gg=1
    #    ws['P1'] = 'Название компании'
    #    ws['P' + str(noom)] = profiles
    #
    #    ws['Q1'] = 'Название компании'
    #    ws['Q' + str(noom)] = profiles
    #
    #    ws['R1'] = 'Название компании'
    #    ws['R' + str(noom)] = profiles
    #
    #    ws['S1'] = 'Название компании'
    #    ws['S' + str(noom)] = profiles
    wb.save(fn)
wb.save(fn)
print('Готово')
wb.close()
input()
